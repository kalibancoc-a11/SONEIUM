from __future__ import annotations

from typing import Optional
from datetime import datetime

from loguru import logger
from openpyxl import Workbook, load_workbook
import os

from openpyxl.worksheet.worksheet import Worksheet
from config import config
from models.account import Account


class Excel:

    def __init__(self, account: Optional[Account] = None, file: Optional[str] = None) -> None:

        self.account = account
        self._file = self._get_file(file)
        self._table = self._get_table()
        self._sheet: Worksheet = self._table.active
        if account:
            self.acc_row = self._find_acc_row(str(self.account.profile_number))

    def change_table(self, table_name: str) -> None:

        self._file = os.path.join(config.PATH_DATA, table_name)
        self._table = self._get_table()

    def connect_account(self, account: Account) -> None:

        self.account = account
        self.acc_row = self._find_acc_row(str(self.account.profile_number))

    def _get_file(self, file: Optional[str]) -> str:

        if not file:
            file = config.PATH_EXCEL
            return file
        file = os.path.join(config.PATH_DATA, file)
        return file

    def _get_table(self) -> Workbook:

        if not os.path.exists(self._file):  # Если файл не существует, создаем его
            table = self._create_excel()
        else:
            table = load_workbook(self._file)
        return table

    def _create_excel(self) -> Workbook:

        table = Workbook()  # Создаем новую таблицу
        table.active['A1'] = 'Profile Number'  # Заполняем ячейки
        if self._file == config.PATH_EXCEL:
            table.active['B1'] = 'Address'  # Заполняем ячейки
            table.active['C1'] = 'Password'  # Заполняем ячейки
            table.active['D1'] = 'Seed'  # Заполняем ячейки
            table.active['E1'] = 'Private Key'  # Заполняем ячейки
            table.active['F1'] = 'Proxy'  # Заполняем ячейки
        table.save(self._file)  # Сохраняем таблицу
        return table

    def _find_acc_row(self, profile_number: str) -> int:

        for row in self._sheet.iter_rows(min_row=2, max_col=1):
            if str(row[0].value) == profile_number:
                return row[0].row
        add_row = self._sheet.max_row + 1
        self._sheet.cell(row=add_row, column=1, value=profile_number)
        self._table.save(self._file)
        return add_row

    def add_row(self, values: list) -> None:

        self._sheet.append(values)
        self._table.save(self._file)

    def set_cell(self, column_name: str, value: str | int | float, row: Optional[int] = None) -> None:

        row = self.acc_row if not row else row

        col_num = self.find_column(column_name)
        self._sheet.cell(row=row, column=col_num, value=value)
        self._table.save(self._file)

    def add_column(self, column_name: str) -> int:

        col_num = self._sheet.max_column + 1
        self._sheet.cell(row=1, column=col_num, value=column_name)
        self._table.save(self._file)
        return col_num

    def find_column(self, column_name: str) -> int:

        for row in self._sheet.iter_rows(max_row=1):
            for cell in row:
                if cell.value == column_name:
                    return cell.column
        logger.warning(f'{self.account.profile_number} Столбец {column_name} не найден, создаем новый.')
        return self.add_column(column_name)

    def get_cell(self, column_name: str, row: Optional[int] = None) -> str | int | None:

        row = self.acc_row if not row else row

        col_num = self.find_column(column_name)

        return self._sheet.cell(row=row, column=col_num).value

    def get_column(self, column_name: str, is_empty_pass: bool = False) -> list[str | int | None]:

        col_num = self.find_column(column_name)
        column_values = []
        for raw in self._sheet.iter_cols(min_col=col_num, max_col=col_num, min_row=2):
            for cell in raw:
                if is_empty_pass and cell.value:
                    column_values.append(cell.value)
                elif not is_empty_pass:
                    column_values.append(cell.value)

        return column_values

    def get_row(self, row: Optional[int] = None) -> list[str | int | None]:

        row = self.acc_row if not row else row
        row_values = []
        for raw in self._sheet.iter_rows(min_row=row, max_row=row):
            for cell in raw:
                row_values.append(cell.value)

        return row_values

    def get_counter(self, column_name: str, row: Optional[int] = None) -> int | float:

        row = self.acc_row if not row else row

        col_num = self.find_column(column_name)
        cell = self._sheet.cell(row=row, column=col_num)

        if cell.value is None:
            cell.value = 0
            self._table.save(self._file)
        elif isinstance(cell.value, str):
            if cell.value.isdigit():
                cell.value = int(cell.value)
                self._table.save(self._file)
            elif cell.value.replace('.', '', 1).isdigit():
                cell.value = float(cell.value)
                self._table.save(self._file)
            else:
                raise TypeError(f'Значение в столбце {column_name} не является числом')

        return cell.value

    def increase_counter(self, column_name: str, number: int = 1, row: Optional[int] = None) -> int:

        row = self.acc_row if not row else row

        col_num = self.find_column(column_name)
        cell = self._sheet.cell(row=row, column=col_num)

        if cell.value is None:
            cell.value = 0
        elif isinstance(cell.value, str):
            if cell.value.isdigit():
                cell.value = int(cell.value)
            else:
                raise TypeError(f'Значение в столбце {column_name} не является числом')

        cell.value += number
        self._table.save(self._file)
        return cell.value

    def set_date(self, column_name: str, row: Optional[int] = None) -> None:

        row = self.acc_row if not row else row

        col_num = self.find_column(column_name)

        self._sheet.cell(row=row, column=col_num, value=datetime.now().strftime(config.date_format))
        self._table.save(self._file)

    def get_date(self, column_name: str, row: Optional[int] = None) -> datetime:

        row = self.acc_row if not row else row

        col_num = self.find_column(column_name)

        date_str = self._sheet.cell(row=row, column=col_num).value
        if date_str:
            date_object = datetime.strptime(date_str, config.date_format)
            return date_object
        logger.error(
            f'{self.account.profile_number} Не нашли дату в столбце {column_name} возвращаем старую дату')
        return datetime.now().replace(year=2000)

    def get_counters(self, column_name: str) -> list[int | float]:

        col_num = self.find_column(column_name)
        column_values = []
        for raw in self._sheet.iter_cols(min_col=col_num, max_col=col_num, min_row=2):
            for cell in raw:
                if cell.value is None:
                    cell.value = 0
                elif isinstance(cell.value, str):
                    if cell.value.isdigit():
                        cell.value = int(cell.value)
                    if cell.value.replace('.', '', 1).isdigit():
                        cell.value = float(cell.value)

                column_values.append(cell.value)
        self._table.save(self._file)
        return column_values
